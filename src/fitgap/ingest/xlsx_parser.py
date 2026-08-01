"""Extract requirements from Excel workbooks using a saved column mapping.

On first run for a given file the mapping is resolved by header auto-detection
or (when running interactively) by prompting the consultant; the result is
saved into fitgap.yaml so subsequent runs are hands-off.
"""

from __future__ import annotations

from pathlib import Path

import typer
from openpyxl import load_workbook

from fitgap.config import XlsxMapping
from fitgap.ingest.docx_parser import (
    AREA_HEADERS,
    ID_HEADERS,
    MIN_REQUIREMENT_LEN,
    PRIORITY_HEADERS,
    TEXT_HEADERS,
    _match_area,
)
from fitgap.models import ParsedRequirement, SourceType


def _read_headers(sheet) -> list[str]:
    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    return [str(v).strip() if v is not None else "" for v in first_row]


def _auto_detect(headers: list[str]) -> XlsxMapping | None:
    def find(candidates: tuple[str, ...]) -> str | None:
        for header in headers:
            cleaned = header.lower()
            if cleaned and any(cleaned == c or c in cleaned for c in candidates):
                return header
        return None

    text = find(TEXT_HEADERS)
    if text is None:
        return None
    return XlsxMapping(
        text=text,
        id=find(ID_HEADERS),
        priority=find(PRIORITY_HEADERS),
        functional_area=find(AREA_HEADERS),
    )


def _prompt_for_mapping(headers: list[str], file_name: str) -> XlsxMapping:
    typer.echo(f"\nNo saved column mapping for {file_name}. Columns found:")
    for idx, header in enumerate(headers, start=1):
        typer.echo(f"  {idx}. {header or '(blank)'}")

    def ask(label: str, required: bool) -> str | None:
        while True:
            raw = typer.prompt(
                f"Which column holds the {label}? (number, or 0 for none)",
                default="0" if not required else None,
            )
            try:
                num = int(raw)
            except ValueError:
                typer.echo("Enter a column number.")
                continue
            if num == 0 and not required:
                return None
            if 1 <= num <= len(headers):
                return headers[num - 1]
            typer.echo("Out of range.")

    return XlsxMapping(
        text=ask("requirement text", required=True),
        id=ask("requirement ID", required=False),
        priority=ask("priority", required=False),
        functional_area=ask("functional area", required=False),
    )


def resolve_mapping(
    path: Path, mapping: XlsxMapping | None, interactive: bool
) -> XlsxMapping:
    """Return a usable mapping: saved > auto-detected > interactive prompt."""
    if mapping is not None:
        return mapping
    workbook = load_workbook(str(path), read_only=True)
    try:
        sheet = workbook.active
        headers = _read_headers(sheet)
    finally:
        workbook.close()
    detected = _auto_detect(headers)
    if detected is not None:
        return detected
    if interactive:
        return _prompt_for_mapping(headers, path.name)
    raise ValueError(
        f"Could not detect a requirement-text column in {path.name} "
        f"(headers: {headers}). Re-run interactively or add a mapping to fitgap.yaml."
    )


def parse_xlsx(path: Path, mapping: XlsxMapping) -> list[ParsedRequirement]:
    workbook = load_workbook(str(path), read_only=True)
    try:
        sheet = workbook[mapping.sheet] if mapping.sheet else workbook.active
        headers = _read_headers(sheet)

        def col(header: str | None) -> int | None:
            if header is None:
                return None
            try:
                return headers.index(header)
            except ValueError:
                return None

        text_col = col(mapping.text)
        if text_col is None:
            raise ValueError(
                f"Mapped column '{mapping.text}' not found in {path.name} "
                f"(headers: {headers})"
            )
        id_col = col(mapping.id)
        priority_col = col(mapping.priority)
        area_col = col(mapping.functional_area)

        results: list[ParsedRequirement] = []
        for row_num, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            def cell(idx: int | None) -> str | None:
                if idx is None or idx >= len(row) or row[idx] is None:
                    return None
                return str(row[idx]).strip() or None

            text = cell(text_col)
            if text is None or len(text) < MIN_REQUIREMENT_LEN:
                continue
            ref = f"{path.name}#row{row_num}"
            source_id = cell(id_col)
            if source_id:
                ref += f" ({source_id})"
            results.append(
                ParsedRequirement(
                    text=text,
                    source=SourceType.XLSX,
                    source_ref=ref,
                    priority=cell(priority_col),
                    functional_area=_match_area(cell(area_col)),
                )
            )
        return results
    finally:
        workbook.close()
