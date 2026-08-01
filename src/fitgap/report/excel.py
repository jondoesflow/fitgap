"""Generate the Excel fit-gap register — the tool's primary deliverable.

Three sheets:

* **Register** — one row per requirement, blank "Consultant review" column,
  live Learn hyperlinks for verified claims, and loud conditional formatting
  for anything the consultant must not miss.
* **Summary** — counts by classification / functional area / verification
  status / effort, with charts.
* **Assumptions & Limitations** — every assumption the model stated, the
  redaction log summary, and generation metadata.

A capability claim (Fit — OOB / Fit — Configuration / Extend) with no verified
live citation is always rendered as "UNCONFIRMED — validate manually" — the
register never shows an unverified claim as settled.
"""

from __future__ import annotations

from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from fitgap import __version__
from fitgap.config import Config
from fitgap.models import (
    CATEGORIES_REQUIRING_VERIFICATION,
    Category,
    Confidence,
    Effort,
    FunctionalArea,
    Requirement,
    SourceReliability,
    VerificationStatus,
    Workspace,
)

UNCONFIRMED_DISPLAY = "UNCONFIRMED — validate manually"

REGISTER_HEADERS = [
    "Req ID",
    "Source",
    "Requirement",
    "Functional area",
    "Classification",
    "Proposed approach",
    "Feature relied on",
    "Learn citation",
    "Preview/deprecated",
    "Confidence",
    "Effort",
    "Assumptions",
    "Consultant review",
    "Notes",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
AMBER_FILL = PatternFill("solid", fgColor="FFE699")
GAP_CUSTOM_FILL = PatternFill("solid", fgColor="F8CBAD")
LOW_CONFIDENCE_FONT = Font(bold=True, color="C00000")
THIN_BORDER = Border(bottom=Side(style="thin", color="D9D9D9"))
WRAP = Alignment(wrap_text=True, vertical="top")

COLUMN_WIDTHS = {
    "A": 10, "B": 26, "C": 55, "D": 18, "E": 22, "F": 55, "G": 34,
    "H": 40, "I": 16, "J": 12, "K": 8, "L": 45, "M": 24, "N": 30,
}


def _citation_cell(req: Requirement) -> tuple[str, str | None]:
    """Return (display value, hyperlink url or None) for the citation column."""
    classification = req.classification
    if classification is None:
        return "", None
    if classification.category not in CATEGORIES_REQUIRING_VERIFICATION:
        return "n/a — no capability claim", None
    verification = req.verification
    if (
        verification is not None
        and verification.status == VerificationStatus.VERIFIED
        and verification.citation_url
    ):
        return verification.page_title or verification.citation_url, verification.citation_url
    return UNCONFIRMED_DISPLAY, None


def _flag_cell(req: Requirement) -> str:
    if req.verification is None:
        return ""
    flags = []
    if req.verification.preview_flag:
        flags.append("PREVIEW")
    if req.verification.deprecated_flag:
        flags.append("DEPRECATED")
    return " + ".join(flags)


def _notes_cell(req: Requirement) -> str:
    notes = []
    if req.source_reliability == SourceReliability.INFERRED:
        notes.append("Inferred from transcript — confirm with client.")
    if req.merged_from:
        notes.append(f"Merged duplicates: {'; '.join(req.merged_from)}")
    if req.verification and req.verification.notes:
        notes.append(req.verification.notes)
    return " ".join(notes)


def _build_register_sheet(sheet: Worksheet, workspace: Workspace) -> None:
    sheet.title = "Register"
    sheet.append(REGISTER_HEADERS)
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for req in workspace.requirements:
        classification = req.classification
        citation_value, citation_url = _citation_cell(req)
        sheet.append(
            [
                req.id,
                req.source_ref,
                req.text,
                req.functional_area.value,
                classification.category.value if classification else "(unclassified)",
                classification.proposed_approach if classification else "",
                classification.feature_relied_on if classification else "",
                citation_value,
                _flag_cell(req),
                classification.confidence.value if classification else "",
                classification.effort.value if classification else "",
                "\n".join(classification.assumptions) if classification else "",
                "",  # Consultant review — deliberately blank
                _notes_cell(req),
            ]
        )
        row = sheet.max_row
        if citation_url:
            cell = sheet.cell(row=row, column=8)
            cell.hyperlink = citation_url
            cell.font = Font(color="0563C1", underline="single")
        for cell in sheet[row]:
            cell.alignment = WRAP
            cell.border = THIN_BORDER

    last_row = sheet.max_row
    data_range = f"A2:N{last_row}"
    if last_row >= 2:
        sheet.conditional_formatting.add(
            data_range,
            FormulaRule(
                formula=[f'$H2="{UNCONFIRMED_DISPLAY}"'], fill=AMBER_FILL
            ),
        )
        sheet.conditional_formatting.add(
            data_range,
            FormulaRule(
                formula=[f'$E2="{Category.GAP_CUSTOM.value}"'], fill=GAP_CUSTOM_FILL
            ),
        )
        sheet.conditional_formatting.add(
            data_range,
            FormulaRule(formula=['$J2="Low"'], font=LOW_CONFIDENCE_FONT),
        )

    for letter, width in COLUMN_WIDTHS.items():
        sheet.column_dimensions[letter].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:N{max(last_row, 2)}"


def _write_block(
    sheet: Worksheet, title: str, rows: list[tuple], start: int
) -> tuple[int, int]:
    """Write a titled two-column block; return (data_first_row, data_last_row)."""
    sheet.cell(row=start, column=1, value=title).font = Font(bold=True, size=12)
    first = start + 1
    for offset, (label, value) in enumerate(rows):
        sheet.cell(row=first + offset, column=1, value=label)
        sheet.cell(row=first + offset, column=2, value=value)
    return first, first + len(rows) - 1


def _build_summary_sheet(sheet: Worksheet, workspace: Workspace) -> None:
    sheet.title = "Summary"
    requirements = workspace.requirements
    classified = [r for r in requirements if r.classification]

    category_counts = Counter(r.classification.category for r in classified)
    category_rows = [(c.value, category_counts.get(c, 0)) for c in Category]
    unclassified = len(requirements) - len(classified)
    if unclassified:
        category_rows.append(("(unclassified)", unclassified))

    area_counts = Counter(r.functional_area for r in requirements)
    area_rows = [
        (a.value, area_counts.get(a, 0)) for a in FunctionalArea if area_counts.get(a)
    ]

    claims = [
        r
        for r in classified
        if r.classification.category in CATEGORIES_REQUIRING_VERIFICATION
    ]
    verified = [
        r
        for r in claims
        if r.verification and r.verification.status == VerificationStatus.VERIFIED
    ]
    pct = round(100 * len(verified) / len(claims), 1) if claims else 0.0
    verification_rows = [
        ("Verified against Microsoft Learn", len(verified)),
        ("UNCONFIRMED — validate manually", len(claims) - len(verified)),
        ("No capability claim (gaps / out of scope)", len(classified) - len(claims)),
    ]

    effort_counts = Counter(r.classification.effort for r in classified)
    effort_rows = [(e.value, effort_counts.get(e, 0)) for e in Effort]
    low_confidence = sum(
        1 for r in classified if r.classification.confidence == Confidence.LOW
    )

    sheet.cell(row=1, column=1, value="Fit-gap summary").font = Font(bold=True, size=14)
    sheet.cell(row=2, column=1, value=f"Total requirements: {len(requirements)}")
    sheet.cell(
        row=3, column=1, value=f"Capability claims verified against live Learn docs: {pct}%"
    )
    sheet.cell(row=4, column=1, value=f"Low-confidence classifications: {low_confidence}")

    cat_first, cat_last = _write_block(sheet, "By classification", category_rows, 6)
    area_start = cat_last + 2
    _write_block(sheet, "By functional area", area_rows, area_start)
    ver_start = area_start + len(area_rows) + 2
    ver_first, ver_last = _write_block(
        sheet, "Verification status", verification_rows, ver_start
    )
    eff_start = ver_last + 2
    _write_block(sheet, "Effort profile (t-shirt)", effort_rows, eff_start)

    bar = BarChart()
    bar.title = "Requirements by classification"
    bar.type = "bar"
    bar.add_data(
        Reference(sheet, min_col=2, min_row=cat_first, max_row=cat_last),
        titles_from_data=False,
    )
    bar.set_categories(Reference(sheet, min_col=1, min_row=cat_first, max_row=cat_last))
    bar.legend = None
    bar.height, bar.width = 9, 16
    sheet.add_chart(bar, "D2")

    pie = PieChart()
    pie.title = "Verification status"
    pie.add_data(
        Reference(sheet, min_col=2, min_row=ver_first, max_row=ver_last),
        titles_from_data=False,
    )
    pie.set_categories(Reference(sheet, min_col=1, min_row=ver_first, max_row=ver_last))
    pie.height, pie.width = 9, 12
    sheet.add_chart(pie, "D22")

    sheet.column_dimensions["A"].width = 42
    sheet.column_dimensions["B"].width = 10


def _build_assumptions_sheet(
    sheet: Worksheet, workspace: Workspace, config: Config
) -> None:
    sheet.title = "Assumptions & Limitations"
    sheet.append(["Req ID", "Classification", "Assumption"])
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for req in workspace.requirements:
        if not req.classification:
            continue
        for assumption in req.classification.assumptions:
            sheet.append([req.id, req.classification.category.value, assumption])

    row = sheet.max_row + 2
    sheet.cell(row=row, column=1, value="Redaction log summary").font = Font(bold=True)
    rule_counts = Counter(e.rule for e in workspace.redaction_log)
    if rule_counts:
        for rule, count in sorted(rule_counts.items()):
            row += 1
            sheet.cell(row=row, column=1, value=rule)
            sheet.cell(row=row, column=2, value=count)
    else:
        row += 1
        sheet.cell(row=row, column=1, value="No redactions recorded.")

    row += 2
    sheet.cell(row=row, column=1, value="Limitations").font = Font(bold=True)
    limitations = [
        "This register is a DRAFT produced by the fitgap tool; the consultant "
        "makes the final call on every row.",
        "Rows marked 'UNCONFIRMED — validate manually' carry a capability claim "
        "that could not be verified against live Microsoft Learn documentation.",
        "Transcript-derived requirements are inferred from conversation and "
        "must be confirmed with the client.",
        "Effort sizes are indicative t-shirt estimates only, dependent on the "
        "assumptions listed above.",
    ]
    for limitation in limitations:
        row += 1
        sheet.cell(row=row, column=1, value=limitation)

    row += 2
    sheet.cell(row=row, column=1, value="Generation metadata").font = Font(bold=True)
    metadata = [
        ("Generated at (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M")),
        ("fitgap version", __version__),
        ("Classification model", config.model),
        ("Workspace created at (UTC)", workspace.created_at.strftime("%Y-%m-%d %H:%M")),
    ]
    for label, value in metadata:
        row += 1
        sheet.cell(row=row, column=1, value=label)
        sheet.cell(row=row, column=2, value=value)

    sheet.column_dimensions["A"].width = 60
    sheet.column_dimensions["B"].width = 24
    sheet.column_dimensions["C"].width = 70


def generate_register(workspace: Workspace, config: Config, path: Path) -> None:
    workbook = Workbook()
    _build_register_sheet(workbook.active, workspace)
    _build_summary_sheet(workbook.create_sheet(), workspace)
    _build_assumptions_sheet(workbook.create_sheet(), workspace, config)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(str(path))
