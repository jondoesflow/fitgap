from datetime import datetime, timezone

from openpyxl import load_workbook

from fitgap.config import Config
from fitgap.models import (
    Category,
    Classification,
    Confidence,
    Effort,
    RedactionEvent,
    Requirement,
    SourceReliability,
    SourceType,
    Verification,
    VerificationStatus,
    Workspace,
)
from fitgap.report import UNCONFIRMED_DISPLAY, generate_register


def classified_req(
    req_id: str,
    category: Category,
    confidence: Confidence = Confidence.HIGH,
    verification: Verification | None = None,
    reliability: SourceReliability = SourceReliability.STATED,
) -> Requirement:
    return Requirement(
        id=req_id,
        text=f"Requirement body for {req_id}.",
        source=SourceType.DOCX,
        source_ref=f"brd.docx#{req_id}",
        source_reliability=reliability,
        classification=Classification(
            category=category,
            proposed_approach="Do the sensible thing.",
            feature_relied_on="Dynamics 365 Sales — some feature",
            effort=Effort.M,
            assumptions=["Assumption one.", "Assumption two."],
            confidence=confidence,
        ),
        verification=verification,
    )


def sample_workspace() -> Workspace:
    return Workspace(
        tool_version="test",
        requirements=[
            classified_req(
                "REQ-001",
                Category.FIT_OOB,
                verification=Verification(
                    status=VerificationStatus.VERIFIED,
                    citation_url="https://learn.microsoft.com/en-us/dynamics365/sales/manage-opportunities",
                    page_title="Manage opportunities",
                    retrieved_at=datetime.now(timezone.utc),
                ),
            ),
            classified_req("REQ-002", Category.FIT_CONFIG),  # claim, unverified
            classified_req("REQ-003", Category.GAP_CUSTOM, confidence=Confidence.LOW),
            classified_req(
                "REQ-004",
                Category.EXTEND_PP,
                verification=Verification(
                    status=VerificationStatus.VERIFIED,
                    citation_url="https://learn.microsoft.com/en-us/power-automate/getting-started",
                    page_title="Get started with Power Automate",
                    preview_flag=True,
                ),
                reliability=SourceReliability.INFERRED,
            ),
            Requirement(
                id="REQ-005",
                text="An unclassified requirement.",
                source=SourceType.XLSX,
                source_ref="backlog.xlsx#row9",
            ),
        ],
        redaction_log=[
            RedactionEvent(rule="client_name", original="Contoso", replacement="[CLIENT]"),
            RedactionEvent(rule="email", original="a@b.com", replacement="[EMAIL]"),
        ],
    )


def make_register(tmp_path):
    path = tmp_path / "register.xlsx"
    generate_register(sample_workspace(), Config(), path)
    return load_workbook(str(path))


def test_sheet_structure(tmp_path):
    workbook = make_register(tmp_path)
    assert workbook.sheetnames == ["Register", "Summary", "Assumptions & Limitations"]
    register = workbook["Register"]
    headers = [c.value for c in register[1]]
    assert headers[:5] == ["Req ID", "Source", "Requirement", "Functional area", "Classification"]
    assert headers[12] == "Consultant review"
    assert register.freeze_panes == "A2"
    assert register.auto_filter.ref is not None


def test_verified_claim_gets_live_hyperlink(tmp_path):
    register = make_register(tmp_path)["Register"]
    cell = register.cell(row=2, column=8)  # REQ-001 citation
    assert cell.value == "Manage opportunities"
    assert cell.hyperlink.target.startswith("https://learn.microsoft.com/")


def test_unverified_claim_reads_unconfirmed(tmp_path):
    register = make_register(tmp_path)["Register"]
    assert register.cell(row=3, column=8).value == UNCONFIRMED_DISPLAY
    assert register.cell(row=3, column=8).hyperlink is None


def test_gap_rows_carry_no_claim(tmp_path):
    register = make_register(tmp_path)["Register"]
    assert "no capability claim" in register.cell(row=4, column=8).value


def test_preview_flag_and_inferred_note(tmp_path):
    register = make_register(tmp_path)["Register"]
    assert register.cell(row=5, column=9).value == "PREVIEW"
    assert "Inferred from transcript" in register.cell(row=5, column=14).value


def test_consultant_review_column_is_blank(tmp_path):
    register = make_register(tmp_path)["Register"]
    for row in range(2, register.max_row + 1):
        assert register.cell(row=row, column=13).value in (None, "")


def test_conditional_formatting_rules_present(tmp_path):
    register = make_register(tmp_path)["Register"]
    formulas = [
        rule.formula[0]
        for cf in register.conditional_formatting
        for rule in cf.rules
    ]
    assert any(UNCONFIRMED_DISPLAY in f for f in formulas)
    assert any(Category.GAP_CUSTOM.value in f for f in formulas)
    assert any('"Low"' in f for f in formulas)


def test_summary_counts_and_charts(tmp_path):
    summary = make_register(tmp_path)["Summary"]
    values = {
        summary.cell(row=r, column=1).value: summary.cell(row=r, column=2).value
        for r in range(1, summary.max_row + 1)
    }
    assert values[Category.FIT_OOB.value] == 1
    assert values[Category.GAP_CUSTOM.value] == 1
    assert values["(unclassified)"] == 1
    assert values["Verified against Microsoft Learn"] == 2
    assert values["UNCONFIRMED — validate manually"] == 1
    assert len(summary._charts) == 2
    header_text = " ".join(
        str(summary.cell(row=r, column=1).value) for r in range(1, 5)
    )
    assert "66.7%" in header_text  # 2 of 3 claims verified


def test_assumptions_sheet_lists_every_assumption(tmp_path):
    sheet = make_register(tmp_path)["Assumptions & Limitations"]
    rows = [
        (sheet.cell(row=r, column=1).value, sheet.cell(row=r, column=3).value)
        for r in range(2, sheet.max_row + 1)
    ]
    req_ids = [r for r, _ in rows if r and str(r).startswith("REQ-")]
    assert len(req_ids) == 8  # 4 classified requirements x 2 assumptions
    col_a = [str(sheet.cell(row=r, column=1).value) for r in range(1, sheet.max_row + 1)]
    assert any("client_name" in v for v in col_a)
    assert any("DRAFT" in v for v in col_a)
